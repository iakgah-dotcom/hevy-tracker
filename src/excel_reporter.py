"""Generate / update the BIG3 Excel report with charts."""
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── colour palette ──────────────────────────────────────────────────────────
COLOURS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "bench":    "2E75B6",
    "deadlift": "C00000",
    "squat":    "70AD47",
    "alt_row":  "EBF3FB",
    "border":   "BFBFBF",
    "title_bg": "D6E4F0",
}

LIFT_COLOURS = {"bench": "2E75B6", "deadlift": "C00000", "squat": "70AD47"}
LIFT_ORDER   = ["bench", "deadlift", "squat"]
LIFT_LABELS  = {"bench": "ベンチプレス", "deadlift": "デッドリフト", "squat": "スクワット"}

OUTPUT_PATH = Path("data/BIG3_tracker.xlsx")


# ── helpers ──────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color=COLOURS["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _header_font(bold: bool = True) -> Font:
    return Font(name="Meiryo", bold=bold, color=COLOURS["header_fg"], size=11)


def _cell_font(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Meiryo", bold=bold, size=size)


def _header_fill(colour: str = COLOURS["header_bg"]) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def _alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOURS["alt_row"])


def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _header_font()
        c.fill = _header_fill()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()


# ── summary sheet ─────────────────────────────────────────────────────────────

def _build_summary(ws, records: list[dict], weekly_data: dict, prs: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"BIG3 週次トラッカー  （更新: {date.today().isoformat()}）"
    c.font = Font(name="Meiryo", bold=True, size=14, color="1F4E79")
    c.fill = PatternFill("solid", fgColor=COLOURS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    # PR table
    pr_headers = ["種目", "推定1RM (kg)", "最大重量 (kg)", "ベストセット日"]
    _write_header_row(ws, 3, pr_headers)
    _set_col_widths(ws, [18, 16, 16, 16, 16, 16, 16, 16])

    for i, lift in enumerate(LIFT_ORDER, 4):
        pr = prs.get(lift, {})
        best = pr.get("best_set")
        ws.cell(row=i, column=1, value=LIFT_LABELS[lift]).font = _cell_font(bold=True)
        ws.cell(row=i, column=2, value=pr.get("max_e1rm", "—"))
        ws.cell(row=i, column=3, value=pr.get("max_weight", "—"))
        ws.cell(row=i, column=4, value=best["date"].isoformat() if best else "—")
        for col in range(1, 5):
            c = ws.cell(row=i, column=col)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                c.fill = _alt_fill()

    # weekly summary table (all weeks, sorted)
    all_weeks: list[str] = sorted(
        {wk for lift in weekly_data for wk in weekly_data[lift]}
    )
    row_start = 8
    ws.merge_cells(f"A{row_start}:H{row_start}")
    t = ws[f"A{row_start}"]
    t.value = "週別サマリー"
    t.font = Font(name="Meiryo", bold=True, size=12, color="1F4E79")
    t.fill = PatternFill("solid", fgColor=COLOURS["title_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_start].height = 22

    summary_headers = [
        "週", "BP e1RM", "BP Volume", "DL e1RM", "DL Volume", "SQ e1RM", "SQ Volume", "合計Volume"
    ]
    _write_header_row(ws, row_start + 1, summary_headers)

    for i, wk in enumerate(reversed(all_weeks)):  # newest first
        r = row_start + 2 + i
        row_data = [wk]
        total_vol = 0.0
        for lift in LIFT_ORDER:
            entry = weekly_data.get(lift, {}).get(wk)
            e1rm = entry["max_e1rm"] if entry else None
            vol  = entry["total_volume"] if entry else None
            row_data += [e1rm, vol]
            if vol:
                total_vol += vol
        row_data.append(total_vol if total_vol else None)

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center")
            c.font = _cell_font()
            if isinstance(val, float):
                c.number_format = "0.0"
            if i % 2 == 0:
                c.fill = _alt_fill()


# ── per-lift sheet ────────────────────────────────────────────────────────────

def _build_lift_sheet(ws, lift: str, records: list[dict], weekly_data: dict) -> None:
    ws.sheet_view.showGridLines = False
    label = LIFT_LABELS[lift]
    colour = LIFT_COLOURS[lift]

    # title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"{label}  詳細ログ"
    c.font = Font(name="Meiryo", bold=True, size=14, color=colour)
    c.fill = PatternFill("solid", fgColor=COLOURS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── set-level log ─────────────────────────────────────────────────────────
    headers = ["日付", "週", "種目名", "重量 (kg)", "回数", "ボリューム", "推定1RM"]
    _write_header_row(ws, 3, headers)
    _set_col_widths(ws, [14, 10, 28, 12, 10, 14, 14, 12])

    lift_records = [r for r in records if r["lift"] == lift]
    lift_records.sort(key=lambda r: r["date"], reverse=True)

    for i, r in enumerate(lift_records, 4):
        row_data = [
            r["date"].isoformat(), f"W{r['week']:02d}",
            r["exercise_name"], r["weight_kg"], r["reps"],
            r["volume"], r["e1rm"],
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center")
            c.font = _cell_font()
            if isinstance(val, float):
                c.number_format = "0.0"
            if i % 2 == 1:
                c.fill = _alt_fill()

    # ── weekly aggregation ────────────────────────────────────────────────────
    all_weeks = sorted(weekly_data.get(lift, {}).keys())
    chart_start_row = max(len(lift_records) + 5, 15)

    ws.merge_cells(f"A{chart_start_row}:G{chart_start_row}")
    t = ws[f"A{chart_start_row}"]
    t.value = "週別集計"
    t.font = Font(name="Meiryo", bold=True, size=12, color=colour)
    t.fill = PatternFill("solid", fgColor=COLOURS["title_bg"])
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[chart_start_row].height = 22

    weekly_headers = ["週", "最大重量 (kg)", "推定1RM (kg)", "総Volume (kg)", "セッション数", "セット数"]
    _write_header_row(ws, chart_start_row + 1, weekly_headers)

    data_start = chart_start_row + 2
    for i, wk in enumerate(all_weeks):
        entry = weekly_data[lift][wk]
        row_data = [
            wk,
            entry["max_weight"],
            entry["max_e1rm"],
            entry["total_volume"],
            entry["sessions"],
            entry["sets"],
        ]
        r = data_start + i
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center")
            c.font = _cell_font()
            if isinstance(val, float):
                c.number_format = "0.0"
            if i % 2 == 0:
                c.fill = _alt_fill()

    # ── chart ─────────────────────────────────────────────────────────────────
    if len(all_weeks) >= 1:
        data_end = data_start + len(all_weeks) - 1
        chart = LineChart()
        chart.title = f"{label} 推定1RM 推移"
        chart.style = 10
        chart.y_axis.title = "推定1RM (kg)"
        chart.x_axis.title = "週"
        chart.width = 22
        chart.height = 12

        # e1rm series (col 3)
        e1rm_data = Reference(ws, min_col=3, max_col=3, min_row=chart_start_row + 1, max_row=data_end)
        chart.add_data(e1rm_data, titles_from_data=True)
        chart.series[0].title = SeriesLabel(v=f"{label} 推定1RM")
        chart.series[0].graphicalProperties.line.solidFill = colour
        chart.series[0].graphicalProperties.line.width = 25000

        cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        chart.set_categories(cats)

        ws.add_chart(chart, f"I{chart_start_row}")


# ── public entry point ────────────────────────────────────────────────────────

def write_report(
    records: list[dict],
    weekly_data: dict,
    prs: dict,
    path: Path = OUTPUT_PATH,
) -> Path:
    """Create or overwrite the Excel report. Returns the output path."""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Summary sheet
    ws_sum = wb.create_sheet("サマリー")
    _build_summary(ws_sum, records, weekly_data, prs)

    # Per-lift sheets
    for lift in LIFT_ORDER:
        ws = wb.create_sheet(LIFT_LABELS[lift])
        _build_lift_sheet(ws, lift, records, weekly_data)

    wb.save(path)
    return path
